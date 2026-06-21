from __future__ import annotations

import csv
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


@dataclass(slots=True)
class PrintJobResult:
    output_folder: Path
    completed_codes: int
    total_copies: int
    missing_codes: list[tuple[str, int]]


@dataclass(slots=True)
class OrderSummary:
    total_codes: int
    total_copies: int
    preview_rows: list[tuple[str, int]]


@dataclass(slots=True)
class ParsedOrders:
    order_counts: dict[str, int]
    remark_ignored_codes: list[tuple[str, int]]
    blank_code_rows: list[dict[str, object]]
    missing_rows: list[dict[str, object]]


def load_order_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        for encoding in ("utf-8-sig", "gbk", "gb18030"):
            try:
                with path.open("r", newline="", encoding=encoding) as file:
                    return [list(row) for row in csv.reader(file)]
            except UnicodeDecodeError:
                continue
        raise ValueError("订单 CSV 编码无法识别，请另存为 UTF-8 或 Excel 后重试。")
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            return [
                ["" if value is None else str(value) for value in row]
                for row in sheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()
    raise ValueError("当前只支持导入 CSV 和 XLSX 订单文件。")


def detect_default_columns(rows: list[list[str]]) -> tuple[int, int]:
    if not rows:
        return 2, 3

    header = [str(value).strip().lower() for value in rows[0]]
    quantity_index = 2 if len(header) > 2 else 0
    code_index = 3 if len(header) > 3 else min(1, max(0, len(header) - 1))

    for index, value in enumerate(header):
        if "数量" in value or "qty" in value or "quantity" in value:
            quantity_index = index
            break

    for index, value in enumerate(header):
        if "编码" in value or "code" in value:
            code_index = index
            break

    return quantity_index, code_index


def parse_order_rows(
    rows: list[list[str]],
    quantity_column: int = 2,
    code_column: int = 3,
    skip_header: bool = False,
) -> ParsedOrders:
    return parse_order_rows_with_remarks(
        rows,
        quantity_column=quantity_column,
        code_column=code_column,
        skip_header=skip_header,
        remark_columns=[],
    )


def parse_order_rows_with_remarks(
    rows: list[list[str]],
    quantity_column: int = 2,
    code_column: int = 3,
    skip_header: bool = False,
    remark_columns: list[int] | None = None,
) -> ParsedOrders:
    order_counts: dict[str, int] = {}
    remark_ignored_codes: list[tuple[str, int]] = []
    blank_code_rows: list[dict[str, object]] = []
    missing_rows: list[dict[str, object]] = []
    remark_columns = remark_columns or []
    data_rows = rows[1:] if skip_header else rows
    found_non_empty_code = False

    for row in data_rows:
        if len(row) <= max(quantity_column, code_column):
            continue

        quantity_text = str(row[quantity_column]).strip()
        try:
            quantity = int(float(quantity_text))
        except ValueError:
            continue
        if quantity <= 0:
            continue

        full_code = str(row[code_column]).strip().upper()
        if not full_code:
            blank_code_rows.append(
                {
                    "row_values": list(row),
                    "reason": "商家编码为空",
                    "full_code": "",
                    "quantity": quantity,
                }
            )
            continue

        found_non_empty_code = True
        if any(len(row) > column and str(row[column]).strip() for column in remark_columns):
            remark_ignored_codes.append((full_code, quantity))
            missing_rows.append(
                {
                    "row_values": list(row),
                    "reason": "备注列不为空，已忽略",
                    "full_code": full_code,
                    "quantity": quantity,
                }
            )
            continue

        order_counts[full_code] = order_counts.get(full_code, 0) + quantity

    if found_non_empty_code:
        missing_rows.extend(blank_code_rows)

    return ParsedOrders(
        order_counts=order_counts,
        remark_ignored_codes=remark_ignored_codes,
        blank_code_rows=blank_code_rows,
        missing_rows=missing_rows,
    )


def summarize_order_counts(order_counts: dict[str, int], preview_limit: int = 20) -> OrderSummary:
    preview_rows = sorted(order_counts.items(), key=lambda item: item[0])[:preview_limit]
    return OrderSummary(
        total_codes=len(order_counts),
        total_copies=sum(order_counts.values()),
        preview_rows=preview_rows,
    )


def build_print_job(
    order_counts: dict[str, int],
    index_rows: list[dict[str, str]],
    output_root: Path,
    folder_name: str,
    forced_missing_codes: list[tuple[str, int]] | None = None,
    source_headers: list[str] | None = None,
    missing_rows: list[dict[str, object]] | None = None,
) -> PrintJobResult:
    target_root = output_root / folder_name
    target_root.mkdir(parents=True, exist_ok=True)

    index_by_code = {
        row.get("full_code", "").strip().upper(): row
        for row in index_rows
        if row.get("full_code")
    }

    completed_codes = 0
    total_copies = 0
    missing_codes: list[tuple[str, int]] = list(forced_missing_codes or [])
    forced_missing_map = {code: quantity for code, quantity in missing_codes}
    missing_rows = list(missing_rows or [])

    for full_code, quantity in order_counts.items():
        if full_code in forced_missing_map:
            continue

        row = index_by_code.get(full_code)
        if row is None:
            missing_codes.append((full_code, quantity))
            missing_rows.append(
                {
                    "row_values": [],
                    "reason": "图库中未找到对应编码",
                    "full_code": full_code,
                    "quantity": quantity,
                }
            )
            continue

        source_path = Path(row.get("output_path", ""))
        if not source_path.exists():
            missing_codes.append((full_code, quantity))
            missing_rows.append(
                {
                    "row_values": [],
                    "reason": "成品图文件不存在",
                    "full_code": full_code,
                    "quantity": quantity,
                }
            )
            continue

        size_folder = f"{row.get('width_cm', '').strip()}-{row.get('height_cm', '').strip()}"
        quantity_folder = f"各{quantity}"
        destination_dir = target_root / size_folder / quantity_folder
        destination_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, destination_dir / source_path.name)
        completed_codes += 1
        total_copies += quantity

    if missing_rows or missing_codes:
        _write_missing_rows_xlsx(
            target_root / "未匹配编码.xlsx",
            source_headers=source_headers or [],
            missing_rows=missing_rows,
        )

    return PrintJobResult(
        output_folder=target_root,
        completed_codes=completed_codes,
        total_copies=total_copies,
        missing_codes=missing_codes,
    )


def _write_missing_rows_xlsx(
    path: Path,
    source_headers: list[str],
    missing_rows: list[dict[str, object]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "未匹配编码"

    max_data_columns = max((len(item.get("row_values", [])) for item in missing_rows), default=0)
    headers = list(source_headers)
    while len(headers) < max_data_columns:
        headers.append(f"第{len(headers) + 1}列")
    headers.extend(["未匹配原因", "图片编码", "数量"])
    sheet.append(headers)

    for item in missing_rows:
        row_values = list(item.get("row_values", []))
        while len(row_values) < max_data_columns:
            row_values.append("")
        row_values.extend(
            [
                str(item.get("reason", "")),
                str(item.get("full_code", "")),
                str(item.get("quantity", "")),
            ]
        )
        sheet.append(row_values)

    workbook.save(path)
